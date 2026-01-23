# ============================================================
#                CONTROL DE ASIGNACIONES - SERMINCO
# ============================================================

from django.db.models import Q
from django.urls import reverse
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User, Group
from django.views.decorators.http import require_POST
import datetime
from django.views.decorators.csrf import csrf_protect
from .models import Expediente, Contrato, OficinaRegional, Mensaje, Anuncio
from django.contrib.auth import authenticate, login, logout
from django.views.decorators.csrf import csrf_protect
from django.contrib import messages
from django.core.paginator import Paginator
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime
from django.utils.dateparse import parse_date
import pandas as pd
import pandas as pd
import plotly.express as px
from datetime import date
from django.http import HttpResponse
from django.db.models import Count
from django.shortcuts import render
from .models import Expediente
from django.utils.dateparse import parse_datetime
from django.utils import timezone
from django.db import transaction
from django.db.models import Exists, OuterRef
from django.views.decorators.csrf import csrf_exempt
from .models import (
    Anuncio,
    AnuncioLectura,
    AnuncioEliminado,  
)


from .forms import (
    CrearUsuarioForm,
)

# ============================================================
#                       FUNCIONES DE ROL
# ============================================================

def has_group(user, group_name):
    """Verifica si el usuario pertenece a un grupo específico."""
    return user.is_authenticated and user.groups.filter(name=group_name).exists()


def user_role(user):
    """Devuelve el rol principal del usuario."""
    if user.is_superuser or has_group(user, "AdministradorLider"):
        return "ADMIN_LIDER"
    if has_group(user, "Administrador"):
        return "ADMIN"
    if has_group(user, "Coordinador"):
        return "COORD"
    if has_group(user, "Supervisor"):
        return "SUP"
    return None


def role_required(groups):
    """Decorador que restringe el acceso según el grupo."""
    def decorator(view_func):
        @login_required
        def wrapper(request, *args, **kwargs):
            user = request.user
            if user.is_superuser or has_group(user, "AdministradorLider"):
                return view_func(request, *args, **kwargs)
            if any(has_group(user, g) for g in groups):
                return view_func(request, *args, **kwargs)
            messages.error(request, "⚠️ No tienes permisos para acceder a esta sección.")
            return redirect("asignaciones:home_router")
        return wrapper
    return decorator

# ============================================================
# AUTENTICACIÓN Y REDIRECCIÓN POR ROL
# ============================================================

@csrf_protect
def login_demo(request):
    """Inicio de sesión personalizado SERMINCO."""
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "").strip()

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, f"👋 Bienvenido, {user.first_name or user.username}.")
            return redirect("asignaciones:home_router")
        else:
            messages.error(request, "❌ Usuario o contraseña incorrectos.")
            return redirect("asignaciones:login")

    if request.user.is_authenticated:
        return redirect("asignaciones:home_router")

    return render(request, "registration/login.html")

@login_required
def home_router(request):
    """Redirige al panel correspondiente según el rol del usuario."""
    user = request.user

    role = user_role(user)
    if role == "ADMIN_LIDER":
        return redirect("asignaciones:admin_lider_menu")
    elif role == "ADMIN":
        return redirect("asignaciones:admin_menu")
    elif role == "COORD":
        return redirect("asignaciones:coordinador_menu")
    elif role == "SUP":
        return redirect("asignaciones:supervisor_panel")
    else:
        messages.warning(request, "Tu cuenta no tiene un rol asignado. Contacta con el AdministradorLider.")
        logout(request)
        return redirect("asignaciones:login")


@csrf_protect
@login_required
def logout_view(request):
    """
    Cierra completamente la sesión del usuario y limpia las cookies.
    """
    try:
        logout(request)  # Cierra la sesión en Django
        request.session.flush()  # Elimina los datos de sesión
        response = redirect("asignaciones:login")
        response.delete_cookie("sessionid")
        response.delete_cookie("csrftoken")
        messages.success(request, "👋 Has cerrado sesión correctamente.")
        return response
    except Exception as e:
        print(f"⚠️ Error al cerrar sesión: {e}")
        messages.error(request, "⚠️ No se pudo cerrar sesión correctamente.")
        return redirect("asignaciones:home_router")

# ============================================================
#                        SUPERVISOR
# ============================================================
from django.contrib.auth.decorators import login_required, user_passes_test

@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=["Supervisor", "AdministradorLider"]).exists())
def supervisor_panel(request):
    """
    Panel principal del Supervisor.
    - Solo los usuarios con rol 'Supervisor' o 'AdministradorLider' pueden acceder.
    - El AdministradorLider tiene control total sobre este panel también.
    """
    return render(request, "roles/supervisor_home.html")

@role_required(["Supervisor"])
def home_supervisor(request):
    return render(request, "roles/supervisor_home.html")

# ============================================================
#            REGISTRO DE VISITA POR SUPERVISOR
# ============================================================
from datetime import datetime
from django.http import JsonResponse
from django.shortcuts import render
from django.contrib import messages
from .models import Expediente, Anuncio
from .decorators import role_required
from django.contrib.auth.models import User

@role_required(["Supervisor"])
def supervisor_registrar(request):
    """
    Permite al supervisor registrar la fecha de visita de un expediente.
    No cambia el estado automáticamente a CONCLUIDO — 
    eso solo ocurre en la vista de 'estado_expediente'.
    """
    supervisor = request.user
    expedientes = Expediente.objects.filter(supervisor=supervisor).select_related("contrato", "oficina")

    # === AJAX: Registro de visita ===
    if request.method == "POST" and request.headers.get("X-Requested-With", "").lower() == "xmlhttprequest":
        try:
            expediente_id = request.POST.get("expediente_id")
            fecha_visita_str = request.POST.get("fecha_visita")

            if not expediente_id or not fecha_visita_str:
                return JsonResponse({"status": "error", "message": "Datos incompletos."})

            # Convertir la fecha de string a objeto date
            try:
                fecha_visita = datetime.strptime(fecha_visita_str, "%Y-%m-%d").date()
            except ValueError:
                return JsonResponse({"status": "error", "message": "Formato de fecha inválido."})

            expediente = Expediente.objects.filter(id=expediente_id, supervisor=supervisor).first()
            if not expediente:
                return JsonResponse({"status": "error", "message": "Expediente no encontrado o no asignado."})

            expediente.fecha_visita = fecha_visita
            expediente.save(update_fields=["fecha_visita"])

            coordinadores = User.objects.filter(groups__name="Coordinador")
            for coord in coordinadores:
                Anuncio.objects.create(
                    titulo=f"Visita registrada - {expediente.siged}",
                    contenido=f"El supervisor {supervisor.get_full_name()} registró la visita el {fecha_visita.strftime('%d/%m/%Y')}.",
                    tipo="INFO",
                    remitente=supervisor,
                    destinatario=coord,
                )

            return JsonResponse({
                "status": "success",
                "message": "Visita registrada correctamente.",
                "fecha_visita": fecha_visita.strftime("%d/%m/%Y"),
            })

        except Exception as e:
            print(f"❌ Error al guardar visita: {e}")
            return JsonResponse({"status": "error", "message": f"Error interno: {str(e)}"})

    # === Render normal ===
    return render(request, "supervisor/registrar.html", {"expedientes": expedientes})

@role_required(["Supervisor", "AdministradorLider"])
def estado_expediente(request):
    """
    Vista unificada para:
    - Buscar expedientes (GET AJAX)
    - Actualizar estado (POST AJAX)
    """

    if request.headers.get("X-Requested-With", "").lower() == "xmlhttprequest" and request.method == "GET":
        siged = request.GET.get("siged", "").strip()
        carta = request.GET.get("carta_linea", "").strip()

        if not siged and not carta:
            return JsonResponse({"status": "error", "message": "Debe ingresar N° SIGED o Carta de línea.", "results": []})

        expedientes = Expediente.objects.all()

        # Supervisores solo ven sus propios expedientes
        if request.user.groups.filter(name="Supervisor").exists():
            expedientes = expedientes.filter(supervisor=request.user)

        # Filtros dinámicos
        filtro = Q()
        if siged:
            filtro |= Q(siged__icontains=siged)
        if carta:
            filtro |= Q(carta_linea__icontains=carta)

        expedientes = (
            expedientes.filter(filtro)
            .select_related("tipo_supervision", "tipo_documento", "oficina")
            .order_by("-fecha_asignacion")[:10]
        )

        resultados = [
            {
                "id": e.id,
                "siged": e.siged or "",
                "carta_linea": e.carta_linea or "",
                "codigo_actividad": e.codigo_actividad or "",
                "tipo_supervision": e.tipo_supervision.nombre if e.tipo_supervision else "",
                "fecha_asignacion": e.fecha_asignacion.strftime("%d/%m/%Y") if e.fecha_asignacion else "",
            }
            for e in expedientes
        ]

        return JsonResponse({"status": "success", "results": resultados})

    # === MODO ACTUALIZACIÓN (POST) ===
    if request.headers.get("X-Requested-With", "").lower() == "xmlhttprequest" and request.method == "POST":
        try:
            exp_id = request.POST.get("expediente")
            fecha_deriv = request.POST.get("fecha_derivacion", "").strip()
            observaciones = request.POST.get("observaciones", "").strip()
            marcado_concluido = request.POST.get("estado", "").upper() == "CONCLUIDO"

            if not exp_id or not fecha_deriv:
                return JsonResponse({"status": "error", "message": "Datos incompletos."})

            expediente = Expediente.objects.filter(id=exp_id).first()
            if not expediente:
                return JsonResponse({"status": "error", "message": "Expediente no encontrado."})

            # Supervisores solo pueden editar los suyos
            if request.user.groups.filter(name="Supervisor").exists():
                if expediente.supervisor != request.user:
                    return JsonResponse({"status": "error", "message": "No autorizado para este expediente."})

            # Guardar cambios
            expediente.fecha_derivacion = fecha_deriv
            expediente.observaciones = observaciones or "-"
            expediente.estado = "CONCLUIDO" if marcado_concluido else "EN PROCESO"
            expediente.save(update_fields=["fecha_derivacion", "observaciones", "estado"])

            return JsonResponse({
                "status": "success",
                "message": f"Expediente {expediente.siged} actualizado correctamente.",
                "expediente": {
                    "id": expediente.id,
                    "siged": expediente.siged,
                    "fecha_derivacion": expediente.fecha_derivacion.strftime("%d/%m/%Y") if hasattr(expediente.fecha_derivacion, "strftime") else expediente.fecha_derivacion,
                    "estado": expediente.estado,
                    "observaciones": expediente.observaciones,
                }
            })

        except Exception as e:
            print("❌ Error en estado_expediente:", e)
            return JsonResponse({"status": "error", "message": f"Error interno: {e}"})

    # === RENDER NORMAL ===
    expedientes = (
        Expediente.objects.filter(supervisor=request.user)
        .select_related("tipo_supervision", "tipo_documento", "oficina")
        .order_by("-fecha_asignacion")
    )

    return render(request, "supervisor/estado.html", {"expedientes": expedientes})


# ============================================================
#                         COORDINADOR
# ============================================================

@role_required(["Coordinador"])
def coordinador_menu(request):
    return render(request, "coordinador/menu.html")

# ============================================================
#          VISTA: REGISTRAR EXPEDIENTE (COORDINADOR)
# ============================================================

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.models import User
from .models import (
    Expediente,
    OficinaRegional,
    Contrato,
    TipoSupervision,
    TipoDocumento,
    Anuncio,
)
from django.db import transaction
from asignaciones.decorators import role_required


@role_required(["Coordinador"])
@transaction.atomic
def coordinador_registrar(request):
    """
    Permite al coordinador registrar un nuevo expediente.
    Incluye selección de contrato, oficina, tipo de supervisión y documento.
    Filtra las listas desplegables con solo los valores válidos definidos.
    """

    # ==== Listas válidas predefinidas ====
    contratos_validos = ["SUP2500192", "SUP2500203", "SUP2500217", "SUP2500235"]
    oficinas_validas = ["Piura", "Lima", "Arequipa", "Moquegua"]
    tipos_documento_validos = ["Informe de Supervisión", "Informe Técnico", "Ficha de Registro", "Resolución"]
    tipos_supervision_validos = [
        "Actos Inseguros", "Atención de Denuncias", "Atención de Solicitud de ITF",
        "Atención de Solicitudes de AVC-AVP", "Comprobación de operaciones",
        "Condiciones de seguridad", "Condiciones de seguridad con observaciones",
        "Control Metrológico CL especial", "Control Volumétrico",
        "Criticidad de Cilindros de GLP", "Denuncia - Actos Inseguros",
        "Denuncia - PRICE", "Denuncias-Informal", "Ejecución/Levantamiento de Med. Seg.",
        "Envasado, Pintado y Canje de cilindros", "Etiquetados de cilindros de GLP",
        "Informalidad", "PRICE", "Por Operaciones",
        "RHO - Inscripción, Modificación",
        "RHO - Modificación de datos, suspensión, cancelación y habilitación",
        "SPIC", "Supervisión Operativa de Seguridad de CD y RD de GLP",
        "Supervisión Operativa de Seguridad de LV GLP",
        "Supervisión Operativa de Seguridad de LVGLP", "Verificación Póliza"
    ]

    # ==== Datos base (para selects) ====
    supervisores = User.objects.filter(groups__name="Supervisor").order_by("first_name")
    contratos = Contrato.objects.filter(numero__in=contratos_validos).order_by("numero")
    oficinas = OficinaRegional.objects.filter(nombre__in=oficinas_validas).order_by("nombre")
    tipos_supervision = TipoSupervision.objects.filter(nombre__in=tipos_supervision_validos).order_by("nombre")
    tipos_documento = TipoDocumento.objects.filter(nombre__in=tipos_documento_validos).order_by("nombre")

    data = {}

    # ==== POST: Registrar expediente ====
    if request.method == "POST":
        data = {k: request.POST.get(k, "").strip() for k in request.POST.keys()}

        campos_obligatorios = ["siged", "carta_linea", "contrato", "oficina", "supervisor_id"]
        faltantes = [c for c in campos_obligatorios if not data.get(c)]

        if faltantes:
            messages.error(request, f"⚠️ Debes completar los campos obligatorios: {', '.join(faltantes)}.")
            return render(request, "coordinador/registrar.html", {
                "data": data,
                "supervisores": supervisores,
                "contratos": contratos,
                "oficinas": oficinas,
                "tipos_supervision_choices": [(t.id, t.nombre) for t in tipos_supervision],
                "tipos_documento_choices": [(t.id, t.nombre) for t in tipos_documento],
            })

        try:
            contrato = get_object_or_404(Contrato, id=data["contrato"])
            oficina = get_object_or_404(OficinaRegional, id=data["oficina"])
            supervisor = get_object_or_404(User, id=data["supervisor_id"])

            tipo_supervision = (
                get_object_or_404(TipoSupervision, id=data["tipo_supervision"])
                if data.get("tipo_supervision") else None
            )
            tipo_documento = (
                get_object_or_404(TipoDocumento, id=data["tipo_documento"])
                if data.get("tipo_documento") else None
            )

            expediente = Expediente.objects.create(
                siged=data["siged"],
                carta_linea=data["carta_linea"],
                codigo=data.get("codigo", ""),
                codigo_actividad=data.get("codigo_actividad", ""),
                razon_social=data.get("razon_social", ""),
                visita_decision=data.get("visita_decision", "NO"),
                fecha_asignacion=data.get("fecha_asignacion") or None,
                contrato=contrato,
                oficina=oficina,
                supervisor=supervisor,
                tipo_supervision=tipo_supervision,
                tipo_documento=tipo_documento,
                estado="EN PROCESO",
            )

            # Crear anuncio informativo
            Anuncio.objects.create(
                titulo=f"Nuevo expediente asignado: {expediente.siged}",
                contenido=f"Has recibido un nuevo expediente asignado por {request.user.get_full_name()}.",
                tipo="INFO",
                destinatario=supervisor,
                remitente=request.user,
            )

            messages.success(request, f"✅ Expediente {expediente.siged} registrado correctamente.")
            return redirect("asignaciones:coordinador_registrar")

        except Exception as e:
            messages.error(request, f"❌ Error al guardar el expediente: {e}")

    # ==== Render inicial o reintento ====
    return render(request, "coordinador/registrar.html", {
        "data": data,
        "supervisores": supervisores,
        "contratos": contratos,
        "oficinas": oficinas,
        "tipos_supervision_choices": [(t.id, t.nombre) for t in tipos_supervision],
        "tipos_documento_choices": [(t.id, t.nombre) for t in tipos_documento],
    })

@role_required(["Coordinador"])
def coordinador_revisar(request):
    """
    Vista para que el Coordinador revise expedientes filtrados por contrato o carta línea.
    Incluye tipo de supervisión y tipo de documento en el contexto para mostrarlos en la tabla.
    """
    # Consulta base optimizada
    qs = (
        Expediente.objects
        .select_related("supervisor", "contrato", "oficina")
        .order_by("-created_at")
    )

    # Filtros de búsqueda
    contrato_id = request.GET.get("contrato")
    carta_linea = request.GET.get("carta_linea", "").strip()
    siged = request.GET.get("siged", "").strip()

    if contrato_id:
        qs = qs.filter(contrato_id=contrato_id)
    if carta_linea:
        qs = qs.filter(carta_linea__icontains=carta_linea)
    if siged:
        qs = qs.filter(siged__icontains=siged)

    # Traer los contratos activos (para el select del filtro)
    contratos = Contrato.objects.select_related("oficina").order_by("numero")

    context = {
        "expedientes": qs,
        "contratos": contratos,
    }
    return render(request, "coordinador/revisar.html", context)



# ============================================================
#                      ADMINISTRADORES
# ============================================================
# Roles:
# - AdministradorLíder → Control total
# - Administrador → Panel simple
# ============================================================

from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render
from .models import Contrato, Expediente

# ============================================================
# UTILIDAD: VALIDADORES DE ROL
# ============================================================

def es_admin_lider(user):
    """Valida si el usuario pertenece al grupo 'AdministradorLider'."""
    return user.is_authenticated and user.groups.filter(name="AdministradorLider").exists()

def es_admin(user):
    """Valida si el usuario pertenece al grupo 'Administrador'."""
    return user.is_authenticated and user.groups.filter(name="Administrador").exists()

# ============================================================
# ADMINISTRADOR LÍDER (control total del sistema)
# ============================================================

@login_required
@user_passes_test(es_admin_lider)
def admin_lider_menu(request):
    """
    Panel principal del AdministradorLíder.
    Tiene acceso completo a todos los paneles (coordinador, supervisor, etc.)
    y puede crear usuarios, revisar y descargar expedientes.
    """
    return render(request, "admin/admin_lider_menu.html")


@login_required
@user_passes_test(es_admin_lider)
def admin_lider_revisar(request):
    """
    Permite revisar expedientes asignados o totales.
    El AdministradorLíder puede filtrar por SIGED o Contrato.
    """
    siged = request.GET.get("siged", "").strip()
    contratos = Contrato.objects.all().order_by("numero")
    expedientes = Expediente.objects.all().order_by("-fecha_asignacion")

    if siged:
        expedientes = expedientes.filter(siged__icontains=siged)

    context = {
        "siged": siged,
        "contratos": contratos,
        "expedientes": expedientes,
    }
    return render(request, "admin/admin_lider_revisar.html", context)

@login_required
@user_passes_test(es_admin_lider)
def admin_lider_descargar(request):
    """
    Vista del Administrador Líder con filtro por contrato y fechas,
    paginación y exportación a Excel con indicador de visita (Sí/No).
    """
    contratos = Contrato.objects.all().order_by("numero")
    contrato_id = request.GET.get("contrato")
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    exportar = request.GET.get("exportar")

    # === QUERY BASE ===
    expedientes = (
        Expediente.objects.all()
        .select_related("contrato", "oficina", "supervisor")
        .order_by("-fecha_asignacion")
    )

    # === FILTROS ===
    if contrato_id:
        expedientes = expedientes.filter(contrato_id=contrato_id)

    if fecha_inicio:
        fecha_i = parse_date(fecha_inicio)
        if fecha_i:
            expedientes = expedientes.filter(fecha_asignacion__gte=fecha_i)

    if fecha_fin:
        fecha_f = parse_date(fecha_fin)
        if fecha_f:
            expedientes = expedientes.filter(fecha_asignacion__lte=fecha_f)

    # === EXPORTAR A EXCEL (rápido sin formato) ===
    if exportar == "1":
        data = [
            {
                "N° SIGED": e.siged,
                "Carta Línea": e.carta_linea,
                "Código OSINERGMIN": e.codigo,
                "Código Actividad": e.codigo_actividad,
                "Razón Social": e.razon_social,
                "Tipo Supervisión": e.tipo_supervision,
                "Tipo Documento": e.tipo_documento,
                "Oficina Regional": e.oficina.nombre if e.oficina else "",
                "Supervisor": e.supervisor.get_full_name() if e.supervisor else "",
                "Fecha Asignación": e.fecha_asignacion.strftime("%d/%m/%Y") if e.fecha_asignacion else "",
                "Fecha Derivación": e.fecha_derivacion.strftime("%d/%m/%Y") if e.fecha_derivacion else "",
                "Visita": "Sí" if e.fecha_visita else "No",
                "Fecha Visita": e.fecha_visita.strftime("%d/%m/%Y") if e.fecha_visita else "",
                "Estado": e.estado,
                "Observaciones": e.observaciones or "",
            }
            for e in expedientes
        ]

        df = pd.DataFrame(data)
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="Expedientes_Admin_Lider.xlsx"'
        df.to_excel(response, index=False)
        return response

    # === PAGINACIÓN ===
    paginator = Paginator(expedientes, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "contratos": contratos,
        "page_obj": page_obj,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "contrato_id": contrato_id,
    }

    return render(request, "admin/admin_lider_descargar.html", context)

@login_required
@user_passes_test(es_admin_lider)
def admin_lider_descargar_excel(request):
    """
    Exporta los expedientes filtrados a un archivo Excel (para Administrador Líder),
    con manejo seguro de valores None, celdas fusionadas y objetos ForeignKey.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.cell.cell import MergedCell
    from datetime import datetime
    from django.http import HttpResponse

    # === Obtener filtros desde la URL ===
    contrato_id = request.GET.get("contrato")
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    # === Base de datos ===
    expedientes = Expediente.objects.all().select_related("contrato", "oficina", "supervisor")

    # === Filtro por contrato ===
    if contrato_id and contrato_id != "None":
        try:
            expedientes = expedientes.filter(contrato_id=int(contrato_id))
        except ValueError:
            pass  # ignora si contrato_id no es válido

    # === Filtro por rango de fechas ===
    if fecha_inicio and fecha_fin and fecha_inicio != "None" and fecha_fin != "None":
        try:
            fecha_i = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            fecha_f = datetime.strptime(fecha_fin, "%Y-%m-%d")
            expedientes = expedientes.filter(fecha_asignacion__range=(fecha_i, fecha_f))
        except ValueError:
            pass

    # === Crear libro Excel ===
    wb = Workbook()
    ws = wb.active
    ws.title = "Expedientes"

    # === Estilos generales ===
    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="004AAD", end_color="004AAD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border_style = Border(
        left=Side(style="thin", color="1E1E1E"),
        right=Side(style="thin", color="1E1E1E"),
        top=Side(style="thin", color="1E1E1E"),
        bottom=Side(style="thin", color="1E1E1E"),
    )

    # === Título del reporte ===
    ws.merge_cells("A1:O1")
    ws["A1"] = "REPORTE DE EXPEDIENTES - ADMINISTRADOR LÍDER"
    ws["A1"].font = title_font
    ws["A1"].fill = PatternFill(start_color="0A2647", end_color="0A2647", fill_type="solid")
    ws["A1"].alignment = center_align

    ws.merge_cells("A2:O2")
    ws["A2"] = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A2"].alignment = center_align

   
    headers = [
        "N° SIGED",
        "Carta Línea",
        "Código OSINERGMIN",
        "Código Actividad",
        "Razón Social",
        "Tipo Supervisión",
        "Tipo Documento",
        "Oficina Regional",
        "Supervisor",
        "Fecha Asignación",
        "Fecha Derivación",
        "Visita",
        "Fecha Visita",
        "Estado",
        "Observaciones",
    ]

    ws.append(headers)

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border_style

  
    row_num = 4
    for e in expedientes:
        visita = "Sí" if e.fecha_visita else "No"

        row_data = [
            e.siged or "",
            e.carta_linea or "",
            e.codigo or "",
            e.codigo_actividad or "",
            e.razon_social or "",
            str(e.tipo_supervision) if e.tipo_supervision else "",
            str(e.tipo_documento) if e.tipo_documento else "",
            e.oficina.nombre if e.oficina else "",
            e.supervisor.get_full_name() if e.supervisor else "",
            e.fecha_asignacion.strftime("%d/%m/%Y") if e.fecha_asignacion else "",
            e.fecha_derivacion.strftime("%d/%m/%Y") if e.fecha_derivacion else "",
            visita,
            e.fecha_visita.strftime("%d/%m/%Y") if e.fecha_visita else "",
            e.estado or "",
            e.observaciones or "",
        ]

        ws.append(row_data)

        for col_num in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col_num).border = border_style
        row_num += 1

    
    for col in ws.columns:
        max_length = 0
        column_letter = None

        for cell in col:
            if isinstance(cell, MergedCell):
                continue  # saltar celdas fusionadas
            if not column_letter:
                column_letter = cell.column_letter
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        if column_letter:
            ws.column_dimensions[column_letter].width = max_length + 3


    filename = f"Expedientes_Lider_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



@role_required(["AdministradorLider"])
@login_required
@transaction.atomic
def admin_lider_catalogos(request):
    """
    Panel de administración de catálogos para el Administrador Líder.
    CRUD de:
      - Contratos
      - Tipos de Supervisión
      - Tipos de Documento
      - Oficinas Regionales
    """

    if request.method == "POST":
        accion = (request.POST.get("accion") or "").strip().lower()
        tipo = (request.POST.get("tipo") or "").strip().lower()

        try:
            # =======================
            # VALIDACIONES BÁSICAS
            # =======================
            if accion not in {"agregar", "editar", "eliminar"}:
                return JsonResponse({
                    "status": "error",
                    "message": f"Acción inválida: {accion}"
                })

            if tipo not in {"contrato", "supervision", "documento", "oficina"}:
                return JsonResponse({
                    "status": "error",
                    "message": f"Tipo inválido: {tipo}"
                })

            # =======================
            # AGREGAR
            # =======================
            if accion == "agregar":
                nombre = (request.POST.get("nombre") or "").strip()
                if not nombre:
                    return JsonResponse({
                        "status": "error",
                        "message": "El nombre no puede estar vacío."
                    })

                if tipo == "contrato":
                    oficina_id = request.POST.get("oficina_id")
                    if not oficina_id:
                        return JsonResponse({
                            "status": "error",
                            "message": "Debe seleccionar una oficina."
                        })
                    oficina = get_object_or_404(OficinaRegional, id=oficina_id)
                    Contrato.objects.create(numero=nombre, oficina=oficina)

                    return JsonResponse({
                        "status": "success",
                        "message": f"Contrato '{nombre}' agregado correctamente."
                    })

                if tipo == "supervision":
                    TipoSupervision.objects.create(nombre=nombre)
                    return JsonResponse({
                        "status": "success",
                        "message": f"Tipo de supervisión '{nombre}' agregado correctamente."
                    })

                if tipo == "documento":
                    TipoDocumento.objects.create(nombre=nombre)
                    return JsonResponse({
                        "status": "success",
                        "message": f"Tipo de documento '{nombre}' agregado correctamente."
                    })

                if tipo == "oficina":
                    OficinaRegional.objects.create(nombre=nombre)
                    return JsonResponse({
                        "status": "success",
                        "message": f"Oficina '{nombre}' agregada correctamente."
                    })

            # =======================
            # ELIMINAR
            # =======================
            if accion == "eliminar":
                item_id = request.POST.get("id")
                if not item_id:
                    return JsonResponse({
                        "status": "error",
                        "message": "ID no proporcionado."
                    })

                modelos = {
                    "contrato": Contrato,
                    "supervision": TipoSupervision,
                    "documento": TipoDocumento,
                    "oficina": OficinaRegional,
                }

                modelos[tipo].objects.filter(id=item_id).delete()

                return JsonResponse({
                    "status": "success",
                    "message": "Registro eliminado correctamente."
                })

            # =======================
            # EDITAR
            # =======================
            if accion == "editar":
                item_id = request.POST.get("id")
                nuevo_nombre = (request.POST.get("nombre") or "").strip()

                if not item_id:
                    return JsonResponse({
                        "status": "error",
                        "message": "ID no proporcionado."
                    })

                if not nuevo_nombre:
                    return JsonResponse({
                        "status": "error",
                        "message": "El nombre no puede estar vacío."
                    })

                if tipo == "contrato":
                    contrato = get_object_or_404(Contrato, id=item_id)
                    contrato.numero = nuevo_nombre

                    oficina_id = request.POST.get("oficina_id")
                    if oficina_id:
                        contrato.oficina = get_object_or_404(OficinaRegional, id=oficina_id)

                    contrato.save()

                    return JsonResponse({
                        "status": "success",
                        "message": "Contrato actualizado correctamente."
                    })

                if tipo == "supervision":
                    TipoSupervision.objects.filter(id=item_id).update(nombre=nuevo_nombre)
                    return JsonResponse({
                        "status": "success",
                        "message": "Tipo de supervisión actualizado correctamente."
                    })

                if tipo == "documento":
                    TipoDocumento.objects.filter(id=item_id).update(nombre=nuevo_nombre)
                    return JsonResponse({
                        "status": "success",
                        "message": "Tipo de documento actualizado correctamente."
                    })

                if tipo == "oficina":
                    OficinaRegional.objects.filter(id=item_id).update(nombre=nuevo_nombre)
                    return JsonResponse({
                        "status": "success",
                        "message": "Oficina actualizada correctamente."
                    })

        except Exception as e:
            return JsonResponse({
                "status": "error",
                "message": f"Ocurrió un error interno: {str(e)}"
            })

    # =======================
    # GET
    # =======================
    context = {
        "contratos": Contrato.objects.select_related("oficina").order_by("numero"),
        "tipos_supervision": TipoSupervision.objects.order_by("nombre"),
        "tipos_documento": TipoDocumento.objects.order_by("nombre"),
        "oficinas": OficinaRegional.objects.order_by("nombre"),
    }

    return render(request, "admin/admin_lider_catalogos.html", context)


# ============================================================
# ADMINISTRADOR (nivel intermedio, acceso restringido)
# ============================================================

@login_required
@user_passes_test(es_admin)
def admin_menu(request):
    """
    Panel principal del Administrador.
    Puede revisar y descargar expedientes propios o asignados,
    pero no puede crear usuarios ni acceder a paneles de otros roles.
    """
    return render(request, "admin/admin_menu.html")

@login_required
@user_passes_test(es_admin)
def admin_revisar(request):
    """
    Vista de revisión de expedientes con filtro por contrato, número SIGED o carta de línea.
    Solo muestra los expedientes asociados al contrato seleccionado.
    """
    # Capturar los parámetros GET
    siged = request.GET.get("siged", "").strip()
    carta_linea = request.GET.get("carta_linea", "").strip()
    contrato_id = request.GET.get("contrato", "").strip()

    # Lista de contratos disponibles
    contratos = Contrato.objects.all().order_by("numero")

    # Query base
    expedientes_qs = Expediente.objects.all().order_by("-fecha_asignacion")

    # === FILTROS ===
    if contrato_id:
        expedientes_qs = expedientes_qs.filter(contrato_id=contrato_id)

    if siged:
        expedientes_qs = expedientes_qs.filter(siged__icontains=siged)

    if carta_linea:
        expedientes_qs = expedientes_qs.filter(carta_linea__icontains=carta_linea)

    # === PAGINACIÓN ===
    paginator = Paginator(expedientes_qs, 15)  # 15 por página
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "siged": siged,
        "carta_linea": carta_linea,
        "contrato_id": contrato_id,
        "contratos": contratos,
        "page_obj": page_obj,
    }

    # Detección AJAX para actualizar solo la tabla (y el contador)
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return render(request, "admin/partials/admin_table.html", context)

    return render(request, "admin/admin_revisar.html", context)

@login_required
@user_passes_test(es_admin)
def admin_descargar(request):
    """
    Vista del Administrador con filtros por contrato y rango de fechas,
    paginación y opción para exportar a Excel con indicador de visita (Sí/No).
    """
    contratos = Contrato.objects.all().order_by("numero")
    contrato_id = request.GET.get("contrato")
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    exportar = request.GET.get("exportar")

    # === QUERY BASE ===
    expedientes = (
        Expediente.objects.all()
        .select_related("contrato", "oficina", "supervisor")
        .order_by("-fecha_asignacion")
    )

    # === FILTROS ===
    if contrato_id:
        expedientes = expedientes.filter(contrato_id=contrato_id)

    if fecha_inicio:
        fecha_i = parse_date(fecha_inicio)
        if fecha_i:
            expedientes = expedientes.filter(fecha_asignacion__gte=fecha_i)

    if fecha_fin:
        fecha_f = parse_date(fecha_fin)
        if fecha_f:
            expedientes = expedientes.filter(fecha_asignacion__lte=fecha_f)

    # === EXPORTAR A EXCEL SIMPLE ===
    if exportar == "1":
        data = [
            {
                "N° SIGED": e.siged,
                "Carta Línea": e.carta_linea,
                "Código OSINERGMIN": e.codigo,
                "Código Actividad": e.codigo_actividad,
                "Razón Social": e.razon_social,
                "Tipo Supervisión": e.tipo_supervision,
                "Tipo Documento": e.tipo_documento,
                "Oficina Regional": e.oficina.nombre if e.oficina else "",
                "Supervisor": e.supervisor.get_full_name() if e.supervisor else "",
                "Fecha Asignación": e.fecha_asignacion.strftime("%d/%m/%Y") if e.fecha_asignacion else "",
                "Fecha Derivación": e.fecha_derivacion.strftime("%d/%m/%Y") if e.fecha_derivacion else "",
                "Visita": "Sí" if e.fecha_visita else "No",
                "Fecha Visita": e.fecha_visita.strftime("%d/%m/%Y") if e.fecha_visita else "",
                "Estado": e.estado,
                "Observaciones": e.observaciones or "",
            }
            for e in expedientes
        ]

        df = pd.DataFrame(data)
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="Expedientes_Admin.xlsx"'
        df.to_excel(response, index=False)
        return response

    # === PAGINACIÓN ===
    paginator = Paginator(expedientes, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "contratos": contratos,
        "page_obj": page_obj,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "contrato_id": contrato_id,
    }

    return render(request, "admin/admin_descargar.html", context)

@login_required
@user_passes_test(es_admin)
def admin_descargar_excel(request):
    """
    Exporta los expedientes filtrados a un archivo Excel con formato profesional.
    """
    contrato_id = request.GET.get("contrato")
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    expedientes = (
        Expediente.objects.all()
        .select_related("contrato", "oficina", "supervisor")
        .order_by("-fecha_asignacion")
    )

    # === FILTROS ===
    if contrato_id:
        expedientes = expedientes.filter(contrato_id=contrato_id)

    if fecha_inicio:
        fecha_i = parse_date(fecha_inicio)
        if fecha_i:
            expedientes = expedientes.filter(fecha_asignacion__gte=fecha_i)

    if fecha_fin:
        fecha_f = parse_date(fecha_fin)
        if fecha_f:
            expedientes = expedientes.filter(fecha_asignacion__lte=fecha_f)

    # === Crear workbook ===
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expedientes"

    # === Estilos ===
    bold_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E4053", end_color="2E4053", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # === Encabezado principal ===
    ws.merge_cells("A1:O1")
    ws["A1"] = "REPORTE DE EXPEDIENTES - SERMINCO"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="0A2647", end_color="0A2647", fill_type="solid")
    ws["A1"].alignment = center_align

    ws.merge_cells("A2:O2")
    ws["A2"] = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A2"].alignment = center_align

    # === Encabezado de tabla ===
    headers = [
        "N° SIGED", "Carta Línea", "Código OSINERGMIN", "Código Actividad",
        "Razón Social", "Tipo Supervisión", "Tipo Documento", "Oficina Regional",
        "Supervisor", "Fecha Asignación", "Fecha Derivación", "Visita", "Fecha Visita",
        "Estado", "Observaciones"
    ]

    ws.append(headers)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # === Cuerpo ===
    for e in expedientes:
        visita = "Sí" if e.fecha_visita else "No"
        ws.append([
            e.siged,
            e.carta_linea,
            e.codigo,
            e.codigo_actividad,
            e.razon_social,
            e.tipo_supervision,
            e.tipo_documento,
            e.oficina.nombre if e.oficina else "",
            e.supervisor.get_full_name() if e.supervisor else "",
            e.fecha_asignacion.strftime("%d/%m/%Y") if e.fecha_asignacion else "",
            e.fecha_derivacion.strftime("%d/%m/%Y") if e.fecha_derivacion else "",
            visita,
            e.fecha_visita.strftime("%d/%m/%Y") if e.fecha_visita else "",
            e.estado,
            e.observaciones or "",
        ])

    # === Borde y formato de cuerpo ===
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border

    # === Ajuste automático de columnas ===
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 3

    # === Exportar archivo ===
    filename = f"Reporte_Expedientes_Admin_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
    wb.save(response)
    return response


# ============================================================
#                        GESTIÓN DE USUARIOS
# ============================================================

@role_required(["AdministradorLider"])
def crear_usuario(request):
    if request.method == "POST":
        form = CrearUsuarioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Usuario creado correctamente.")
            return redirect("asignaciones:lista_usuarios")
        else:
            messages.error(request, "⚠️ Revisa los campos marcados.")
    else:
        form = CrearUsuarioForm()

    return render(
        request,
        "usuarios/crear.html",
        {
            "form": form,
            "modo": "crear",
            "titulo": "Crear Usuario"
        }
    )


@role_required(["AdministradorLider"])
def lista_usuarios(request):
    usuarios = User.objects.all().order_by("id")
    usuarios_data = [
        {
            "id": u.id,
            "username": u.username,
            "first_name": u.first_name,
            "last_name": u.last_name,
            "email": u.email,
            "group": u.groups.first().name if u.groups.exists() else "Sin rol",
        }
        for u in usuarios
    ]
    return render(request, "asignaciones/lista_usuarios.html", {"usuarios": usuarios_data})

@role_required(["AdministradorLider"])
def eliminar_usuario(request, user_id):
    usuario = get_object_or_404(User, id=user_id)
    if usuario.is_superuser or usuario.username.lower() == "admin":
        messages.error(request, "⚠️ No se puede eliminar al administrador principal.")
    else:
        usuario.delete()
        messages.success(request, f"✅ Usuario '{usuario.username}' eliminado correctamente.")
    return redirect("asignaciones:lista_usuarios")

# ============================================================
#        ADMINISTRADOR LÍDER - EDITAR USUARIO
# ============================================================

from django.contrib.auth.models import User, Group
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .forms import CrearUsuarioForm as UsuarioForm

@role_required(["AdministradorLider"])
def editar_usuario(request, user_id):
    usuario = get_object_or_404(User, id=user_id)

    if request.method == "POST":
        form = CrearUsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(
                request,
                f"✅ Usuario '{usuario.username}' actualizado correctamente."
            )
            return redirect("asignaciones:lista_usuarios")
        else:
            messages.error(request, "⚠️ Revisa los campos marcados.")
    else:
        form = CrearUsuarioForm(instance=usuario)

    return render(
        request,
        "usuarios/editar_usuario.html",
        {
            "form": form,
            "modo": "editar",
            "titulo": "Editar Usuario"
        }
    )

# ============================================================
#                       BANDEJA 
# ============================================================

@login_required
def bandeja(request):
    tipo = request.GET.get("tipo", "recibidos")

    # =========================
    # ENVIAR MENSAJE
    # =========================
    if request.method == "POST":
        destinatario_username = request.POST.get("destinatario")
        asunto = request.POST.get("asunto")
        cuerpo = request.POST.get("contenido")

        if not destinatario_username or not asunto or not cuerpo:
            messages.error(request, "Completa todos los campos")
            return redirect(f"{reverse('asignaciones:bandeja')}?tipo=enviados")

        try:
            destinatario = User.objects.get(username=destinatario_username)
        except User.DoesNotExist:
            messages.error(request, "El usuario destinatario no existe")
            return redirect(f"{reverse('asignaciones:bandeja')}?tipo=enviados")

        Mensaje.objects.create(
            remitente=request.user,
            destinatario=destinatario,
            asunto=asunto,
            cuerpo=cuerpo
        )

        messages.success(request, "Mensaje enviado correctamente")
        # 🔹 Redirige siempre a "enviados" después de enviar
        return redirect(f"{reverse('asignaciones:bandeja')}?tipo=enviados")

    # =========================
    # LISTADO DE MENSAJES
    # =========================
    if tipo == "enviados":
        mensajes = Mensaje.objects.filter(
            remitente=request.user,
            eliminado_por_remitente=False
        )
    else:
        mensajes = Mensaje.objects.filter(
            destinatario=request.user,
            eliminado_por_destinatario=False
        )

    return render(request, "misc/bandeja.html", {
        "mensajes": mensajes,
        "tipo": tipo
    })


@login_required
def leer_mensaje(request, pk):
    mensaje = get_object_or_404(Mensaje, pk=pk)

    # 🔐 Seguridad total
    if request.user not in [mensaje.remitente, mensaje.destinatario]:
        return JsonResponse({"error": "No autorizado"}, status=403)

    if request.user == mensaje.destinatario and not mensaje.leido:
        mensaje.leido = True
        mensaje.save(update_fields=["leido"])

    return JsonResponse({
        "id": mensaje.id,
        "asunto": mensaje.asunto,
        "cuerpo": mensaje.cuerpo,
        "remitente": mensaje.remitente.username if mensaje.remitente else "Sistema",
        "fecha": mensaje.creado_en.strftime("%d/%m/%Y %H:%M")
    })


@login_required
def eliminar_mensaje(request, pk):
    mensaje = get_object_or_404(Mensaje, pk=pk)

    if request.user == mensaje.remitente:
        mensaje.eliminado_por_remitente = True
    elif request.user == mensaje.destinatario:
        mensaje.eliminado_por_destinatario = True
    else:
        return JsonResponse({"error": "No autorizado"}, status=403)

    mensaje.save(update_fields=[
        "eliminado_por_remitente",
        "eliminado_por_destinatario"
    ])

    messages.success(request, "Mensaje eliminado")
    return redirect("asignaciones:bandeja")


@login_required
def mensajes_parciales(request, tipo):
    """
    Devuelve solo los rows HTML de los mensajes según tipo (recibidos/enviados)
    """
    if tipo == "enviados":
        mensajes = Mensaje.objects.filter(
            remitente=request.user,
            eliminado_por_remitente=False
        )
    else:
        mensajes = Mensaje.objects.filter(
            destinatario=request.user,
            eliminado_por_destinatario=False
        )

    return render(request, "misc/mensajes_parciales.html", {
        "mensajes": mensajes
    })


@csrf_exempt
@login_required
def eliminar_mensajes_ajax(request):
    if request.method == "POST":
        data = json.loads(request.body)
        ids = data.get("ids", [])

        for m_id in ids:
            try:
                mensaje = Mensaje.objects.get(pk=m_id)
                if request.user == mensaje.remitente:
                    mensaje.eliminado_por_remitente = True
                elif request.user == mensaje.destinatario:
                    mensaje.eliminado_por_destinatario = True
                mensaje.save()
            except Mensaje.DoesNotExist:
                continue

        return JsonResponse({"success": True})
    return JsonResponse({"error": "Método no permitido"}, status=405)


# ============================================================
#                  ANUNCIOS - CREACIÓN Y LISTADO
# ============================================================

from django.contrib.auth.models import Group, User
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Anuncio
from django.db import models

@login_required
def anuncios(request):
    user = request.user
    ahora = timezone.now()

    puede_crear = user.groups.filter(name__in=["Administrador", "AdministradorLider"]).exists()
    puede_ver_metricas = puede_crear

    base_qs = (
        Anuncio.objects
        .exclude(eliminaciones__usuario=user)
        .filter(
            Q(fecha_inicio__lte=ahora) | Q(fecha_inicio__isnull=True),
            Q(fecha_fin__gte=ahora) | Q(fecha_fin__isnull=True),
        )
        .annotate(
            leido=Exists(
                AnuncioLectura.objects.filter(
                    anuncio=OuterRef("pk"),
                    usuario=user
                )
            )
        )
    )

    if user.groups.filter(name="AdministradorLider").exists():
        anuncios_qs = base_qs.filter(remitente=user)

    elif user.groups.filter(name="Administrador").exists():
        anuncios_qs = base_qs.filter(
            Q(remitente=user)
            | Q(destinatario=user)
            | Q(grupo_destino__in=user.groups.all())
            | Q(tipo="general")
        ).distinct()

    else:
        anuncios_qs = base_qs.filter(
            Q(destinatario=user)
            | Q(grupo_destino__in=user.groups.all())
            | Q(tipo="general")
        )

    # 🔒 CONTADOR (NO TOCADO)
    contador_anuncios = anuncios_qs.exclude(
        lecturas__usuario=user
    ).count()

    return render(request, "misc/anuncios.html", {
        "anuncios": anuncios_qs.order_by("-fecha_creacion"),
        "contador_anuncios": contador_anuncios,
        "puede_crear": puede_crear,
        "puede_ver_metricas": puede_ver_metricas,
    })



@login_required
def crear_anuncio(request):
    user = request.user

    if not user.groups.filter(name__in=["AdministradorLider", "Administrador"]).exists():
        messages.error(request, "❌ No tienes permiso para crear anuncios.")
        return redirect("asignaciones:anuncios")

    grupos = Group.objects.all().order_by("name")
    usuarios = User.objects.all().order_by("first_name", "last_name")

    if request.method == "POST":
        titulo = request.POST.get("titulo", "").strip()
        contenido = request.POST.get("contenido", "").strip()
        tipo = request.POST.get("tipo", "general")
        grupo_id = request.POST.get("grupo_destino") or None
        destinatario_id = request.POST.get("destinatario") or None

        fecha_inicio_raw = request.POST.get("fecha_inicio")
        fecha_fin_raw = request.POST.get("fecha_fin")
        fecha_inicio = parse_datetime(fecha_inicio_raw) if fecha_inicio_raw else None
        fecha_fin = parse_datetime(fecha_fin_raw) if fecha_fin_raw else None

        if not titulo or not contenido:
            messages.error(request, "Todos los campos obligatorios deben completarse.")
            return redirect("asignaciones:crear_anuncio")

        if tipo == "grupo" and not grupo_id:
            messages.error(request, "Debes seleccionar un grupo destino.")
            return redirect("asignaciones:crear_anuncio")

        if tipo == "individual" and not destinatario_id:
            messages.error(request, "Debes seleccionar un usuario destinatario.")
            return redirect("asignaciones:crear_anuncio")

        if fecha_inicio and fecha_fin and fecha_fin < fecha_inicio:
            messages.error(request, "La fecha de fin no puede ser menor que la fecha de inicio.")
            return redirect("asignaciones:crear_anuncio")

        if fecha_inicio and fecha_inicio < timezone.now():
            messages.error(request, "La fecha de inicio no puede ser en el pasado.")
            return redirect("asignaciones:crear_anuncio")

        Anuncio.objects.create(
            titulo=titulo,
            contenido=contenido,
            tipo=tipo,
            remitente=user,
            grupo_destino_id=grupo_id if tipo == "grupo" else None,
            destinatario_id=destinatario_id if tipo == "individual" else None,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
        )

        messages.success(request, "✅ Anuncio publicado correctamente.")
        return redirect("asignaciones:anuncios")

    return render(request, "misc/crear_anuncio.html", {
        "grupos": grupos,
        "usuarios": usuarios,
    })



@login_required
@transaction.atomic
def marcar_anuncio_leido(request, pk):
    if request.method != "POST":
        return JsonResponse({"status": "error"}, status=405)

    anuncio = get_object_or_404(Anuncio, pk=pk)

    AnuncioLectura.objects.get_or_create(
        anuncio=anuncio,
        usuario=request.user
    )

    return JsonResponse({
        "status": "ok",
        "no_leidos": contar_anuncios_no_leidos(request.user)
    })



@login_required
def eliminar_anuncio(request, pk):
    if request.method != "POST":
        return JsonResponse({"status": "error"}, status=405)

    anuncio = get_object_or_404(Anuncio, pk=pk)

    AnuncioEliminado.objects.get_or_create(
        anuncio=anuncio,
        usuario=request.user
    )

    return JsonResponse({
        "status": "ok",
        "no_leidos": contar_anuncios_no_leidos(request.user)
    })



@login_required
def anuncios_no_leidos(request):
    return JsonResponse({
        "count": contar_anuncios_no_leidos(request.user)
    })


@login_required
def anuncios_metricas(request):
    if not request.user.groups.filter(name__in=["Administrador", "AdministradorLider"]).exists():
        return redirect("asignaciones:anuncios")

    anuncios_qs = (
        Anuncio.objects
        .select_related("remitente")
        .prefetch_related("lecturas__usuario")  
        .order_by("-fecha_creacion")
    )

    return render(request, "misc/anuncios_metricas.html", {
        "anuncios": anuncios_qs
    })


# ============================================================
#                 FUNCIONES AUXILIARES
# ============================================================

def contar_anuncios_no_leidos(user):
    ahora = timezone.now()

    return (
        Anuncio.objects
        .exclude(lecturas__usuario=user)         
        .exclude(eliminaciones__usuario=user)   
        .filter(
            Q(destinatario=user)
            | Q(grupo_destino__in=user.groups.all())
            | Q(tipo="general"),
            Q(fecha_inicio__lte=ahora) | Q(fecha_inicio__isnull=True),
            Q(fecha_fin__gte=ahora) | Q(fecha_fin__isnull=True),
        )
        .distinct()
        .count()
    )

from django.shortcuts import render
from django.db import models
import pandas as pd
import plotly.express as px
from asignaciones.models import Expediente 

# ============================================================
#                       REPORTES
# ============================================================
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import Expediente
from django.db import models

@login_required
@role_required(["Administrador", "AdministradorLider"])
def reportes(request):
    """
    Renderiza la página de reportes con los totales visibles
    y prepara los datos para los gráficos en Chart.js vía AJAX.
    """

    # Datos totales
    total_expedientes = Expediente.objects.count()
    total_oficinas = Expediente.objects.values('oficina').distinct().count()
    total_supervisores = Expediente.objects.values('supervisor').distinct().count()

    context = {
        "total_expedientes": total_expedientes,
        "total_oficinas": total_oficinas,
        "total_supervisores": total_supervisores,
    }

    return render(request, "admin/reportes_dashboard.html", context)

@login_required
@role_required(["Administrador", "AdministradorLider"])
def reportes_json(request):
    # Datos para gráficos
    data_contratos = Expediente.objects.values('contrato__numero').annotate(total=models.Count('id')).order_by('contrato__numero')
    data_oficinas = Expediente.objects.values('oficina__nombre').annotate(total=models.Count('id')).order_by('oficina__nombre')
    data_supervisores = Expediente.objects.values('supervisor__username').annotate(total=models.Count('id')).order_by('supervisor__username')

    return JsonResponse({
        "labels_contratos": [d['contrato__numero'] for d in data_contratos],
        "values_contratos": [d['total'] for d in data_contratos],
        "labels_oficinas": [d['oficina__nombre'] for d in data_oficinas],
        "values_oficinas": [d['total'] for d in data_oficinas],
        "labels_supervisores": [d['supervisor__username'] for d in data_supervisores],
        "values_supervisores": [d['total'] for d in data_supervisores],
    })



# ============================================================
#                 EXPORTAR REPORTE A EXCEL
# ============================================================
def exportar_excel(request):
    fecha_inicio = request.GET.get("fecha_inicio", "2025-01-01")
    fecha_fin = request.GET.get("fecha_fin", str(date.today()))

    queryset = Expediente.objects.filter(fecha_asignacion__range=[fecha_inicio, fecha_fin]).values(
        "siged", "contrato__numero", "oficina__nombre", "supervisor__username",
        "fecha_asignacion", "estado"
    )
    df = pd.DataFrame(list(queryset))
    df.rename(columns={
        "siged": "N° SIGED",
        "contrato__numero": "Contrato",
        "oficina__nombre": "Oficina",
        "supervisor__username": "Supervisor",
        "fecha_asignacion": "Fecha de Asignación",
        "estado": "Estado"
    }, inplace=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_expedientes_{fecha_inicio}_a_{fecha_fin}.xlsx"'
    df.to_excel(response, index=False)
    return response

# ============================================================
#                ENDPOINT AJAX - AUTOCOMPLETADO (MEJORADO)
# ============================================================

from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import Expediente
from .decorators import role_required


@role_required(["AdministradorLider", "Administrador", "Coordinador", "Supervisor"])
@login_required
def ajax_autocomplete(request):
    """
    Retorna coincidencias para autocompletar campos (SIGED o Carta Línea)
    mostrando solo los expedientes visibles por el usuario actual.
    - Los supervisores solo ven sus expedientes asignados.
    - Los administradores y coordinadores pueden ver todos.
    """
    query = request.GET.get("q", "").strip()
    tipo = request.GET.get("tipo", "").strip()  # puede ser "siged" o "carta_linea"

    if not query or not tipo:
        return JsonResponse({"results": []})

    # Filtrar base según el rol del usuario
    user = request.user
    expedientes = Expediente.objects.all()

    if user.groups.filter(name="Supervisor").exists():
        expedientes = expedientes.filter(supervisor=user)

    # Filtro según el campo de búsqueda
    if tipo == "siged":
        expedientes = expedientes.filter(siged__icontains=query)
    elif tipo == "carta_linea":
        expedientes = expedientes.filter(carta_linea__icontains=query)
    else:
        return JsonResponse({"results": []})

    # Armar lista JSON detallada para la tabla del frontend
    resultados = [
        {
            "id": e.id,
            "siged": e.siged,
            "carta_linea": e.carta_linea,
            "razon_social": e.razon_social or "",
            "contrato": e.contrato.numero if e.contrato else "",
            "oficina": e.oficina.nombre if e.oficina else "",
            "fecha_asignacion": e.fecha_asignacion.strftime("%d/%m/%Y") if e.fecha_asignacion else "",
        }
        for e in expedientes[:10]
    ]

    return JsonResponse({"results": resultados})


# ============================================================
#        ENDPOINT AJAX - REGISTRAR VISITA (Supervisor)
# ============================================================

from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404
from .models import Expediente

@require_POST
@role_required(["Supervisor", "AdministradorLider"])
def ajax_registrar_visita(request):
    """
    Permite al Supervisor registrar o actualizar la fecha de visita
    desde una llamada AJAX sin recargar la página.
    """
    try:
        expediente_id = request.POST.get("expediente_id")
        fecha_visita = request.POST.get("fecha_visita")

        expediente = get_object_or_404(Expediente, id=expediente_id)
        expediente.fecha_visita = fecha_visita
        expediente.save()

        return JsonResponse({"success": True, "message": "Visita registrada correctamente."})
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)})

# ============================================================
#     ENDPOINT AJAX - ACTUALIZAR ESTADO DE EXPEDIENTE
# ============================================================
from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from .models import Expediente

@require_POST
@role_required(["AdministradorLider", "Administrador", "Coordinador", "Supervisor"])
def ajax_actualizar_estado(request):
    """
    Permite actualizar el estado de un expediente vía AJAX.
    Usado por roles con permisos para modificar (Coordinador, Admins, Supervisor).
    """
    try:
        expediente_id = request.POST.get("expediente_id")
        nuevo_estado = request.POST.get("nuevo_estado")

        expediente = get_object_or_404(Expediente, id=expediente_id)
        expediente.estado = nuevo_estado
        expediente.save()

        return JsonResponse({
            "success": True,
            "message": f"El estado del expediente {expediente.siged} fue actualizado a '{nuevo_estado}'."
        })

    except Exception as e:
        return JsonResponse({
            "success": False,
            "message": f"Error al actualizar estado: {str(e)}"
        })


